import importlib.util
import mimetypes
from collections.abc import Callable, Collection, Iterable, Sequence
from numbers import Number
from pathlib import Path
from threading import Lock
from typing import Any, cast

# noinspection PyPackageRequirements
import numpy as np
import pandas as pd
import pyqtgraph as pg  # type: ignore

# noinspection PyPackageRequirements
from numpy.typing import NDArray
from pyqtgraph import GraphicsScene, PlotWidget
from pyqtgraph.GraphicsScene.mouseEvents import MouseClickEvent  # type: ignore
from pyqtgraph.exporters import ImageExporter
from qtpy.QtCore import (
    QCoreApplication,
    QPointF,
    QRectF,
    Qt,
    Slot,
)
from qtpy.QtGui import (
    QAction,
    QBrush,
    QCloseEvent,
    QFont,
    QGuiApplication,
    QPen,
    QScreen,
    QShowEvent,
)
from qtpy.QtWidgets import QMainWindow, QMessageBox, QWidget

from ..plot_data_item import PlotDataItem
from ..utils import (
    DataMode,
    HeaderWithUnit,
    SpectrometerData,
    copy_to_clipboard,
    load_data,
    p_tag,
    tag,
    the,
)
from ..widgets.preferences import Preferences
from .gui.frequency_domain_gui import FrequencyDomainGUI

__all__ = ["FrequencyDomainWindow"]


_translate = QCoreApplication.translate


class FrequencyDomainWindow(FrequencyDomainGUI):
    supported_modes: Collection[DataMode] = (
        DataMode.FS,
        DataMode.PSK,
        DataMode.PSK_WITH_JUMP,
    )

    def __init__(
        self,
        file_path: Path | None = None,
        parent: QWidget | None = None,
        flags: Qt.WindowType = Qt.WindowType.Window,
    ) -> None:
        super().__init__(parent, flags)

        self._data_mode: DataMode = DataMode.unknown

        self._ghost_line: pg.PlotDataItem = self.figure.plot(np.empty(0), name="")
        self._plot_line: pg.PlotDataItem = self.figure.plot(np.empty(0), name="")
        self._ghost_data: PlotDataItem = PlotDataItem()
        self._plot_data: PlotDataItem = PlotDataItem()

        self._ignore_scale_change: Lock = Lock()

        self.user_found_lines: pg.PlotDataItem = self._canvas.scatterPlot(
            np.empty(0), symbol="o", pxMode=True
        )
        self.automatically_found_lines: pg.PlotDataItem = self._canvas.scatterPlot(
            np.empty(0), symbol="o", pxMode=True
        )
        self.user_found_lines_data: NDArray[np.float64] = np.empty(0)

        self._cursor_balloon: pg.TextItem = pg.TextItem()
        self.figure.addItem(self._cursor_balloon)

        self._mouse_moved_signal_proxy: pg.SignalProxy = pg.SignalProxy(
            cast(GraphicsScene, self.figure.scene()).sigMouseMoved,
            rateLimit=10,
            slot=self.on_mouse_moved,
        )
        self._axis_range_changed_signal_proxy: pg.SignalProxy = pg.SignalProxy(
            self.figure.sigRangeChanged, rateLimit=20, slot=self.on_lim_changed
        )

        self.setup_ui()
        self._setup_colors()

        self.load_config()

        self.setup_ui_actions()

        if file_path is not None and file_path.exists():
            loaded: bool = self.load_data(file_path)
            self.toolbar.load_trace_action.setEnabled(loaded)
            if loaded:
                self.set_config_value("open", "location", file_path.parent)
                self.load_catalog()

    def setup_ui(self) -> None:
        self.hide_cursors()

        self.set_plot_line_appearance()
        self.set_axis_line_appearance()
        self.set_marks_appearance()
        self.set_crosshair_lines_appearance()

        with the(self._canvas) as canvas:
            # customize menu
            titles_to_leave: list[str] = [
                canvas.ctrl.alphaGroup.parent().title(),
                canvas.ctrl.gridGroup.parent().title(),
            ]
            action: QAction
            for action in canvas.ctrlMenu.actions():
                if action.text() not in titles_to_leave:
                    canvas.ctrlMenu.removeAction(action)
            canvas.vb.menu = canvas.ctrlMenu
            canvas.ctrlMenu = None
            canvas.getViewBox().getMenu(canvas).addAction(self._view_all_action)
            canvas.ctrl.autoAlphaCheck.setChecked(False)
            canvas.ctrl.autoAlphaCheck.hide()
        self.figure.sceneObj.contextMenu = None

        self._install_translation()

    def closeEvent(self, event: QCloseEvent) -> None:
        close_code: int
        if self._data_mode == DataMode.unknown:  # nothing is loaded
            close_code = QMessageBox.StandardButton.Yes
        else:
            # senseless joke in the loop
            close: QMessageBox = QMessageBox(self)
            close.setText(self.tr("Are you sure?"))
            close.setIcon(QMessageBox.Icon.Question)
            close.setWindowIcon(self.windowIcon())
            close.setWindowTitle(self.tr("Spectrometer Data Viewer"))
            close.setStandardButtons(
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No
                | QMessageBox.StandardButton.Cancel
            )
            close_code = QMessageBox.StandardButton.No
            while close_code == QMessageBox.StandardButton.No:
                close_code = close.exec()

        if close_code == QMessageBox.StandardButton.Yes:
            self.settings.save(self)
            self.settings.sync()

            from .. import windows

            windows.remove(self)

            event.accept()
        elif close_code == QMessageBox.StandardButton.Cancel:
            event.ignore()

    def showEvent(self, event: QShowEvent) -> None:

        from .. import windows

        windows.append(self)

        window: QMainWindow
        while windows:
            for window in windows:
                if window.isHidden():
                    window.close()
                    break
            else:
                break

        event.accept()

    def load_config(self) -> None:
        with self._loading:
            # Fallback: Center the window
            screen: QScreen = QGuiApplication.primaryScreen()
            self.move(
                round(0.5 * (screen.size().width() - self.size().width())),
                round(0.5 * (screen.size().height() - self.size().height())),
            )

            self.settings.restore(self)

            self.box_frequency.load_config()
            self.box_voltage.load_config()
            self.box_find_lines.load_config()
            self.box_found_lines.load_config()

            if (
                self.get_config_value("display", "unit", PlotDataItem.VOLTAGE_DATA, str)
                == PlotDataItem.GAMMA_DATA
            ):
                self._plot_data.y_data_type = PlotDataItem.GAMMA_DATA
                self.box_find_lines.set_data_type(PlotDataItem.GAMMA_DATA)
            else:
                self._plot_data.y_data_type = PlotDataItem.VOLTAGE_DATA
                self.box_find_lines.set_data_type(PlotDataItem.VOLTAGE_DATA)
            self.display_gamma_or_voltage()

    def setup_ui_actions(self) -> None:
        self.toolbar.open_action.triggered.connect(self.on_open_action_triggered)
        self.toolbar.clear_action.triggered.connect(self.on_clear_action_triggered)
        self.toolbar.open_ghost_action.triggered.connect(
            self.on_open_ghost_action_triggered
        )
        self.toolbar.clear_ghost_action.triggered.connect(
            self.on_clear_ghost_action_triggered
        )
        self.toolbar.differentiate_action.toggled.connect(
            self.on_differentiate_action_toggled
        )
        self.toolbar.save_data_action.triggered.connect(self.on_save_data_triggered)
        self.toolbar.copy_figure_action.triggered.connect(self.on_copy_figure_triggered)
        self.toolbar.save_figure_action.triggered.connect(self.on_save_figure_triggered)
        self.toolbar.load_trace_action.triggered.connect(
            self.on_load_found_lines_triggered
        )
        self.toolbar.copy_trace_action.triggered.connect(
            self.on_copy_found_lines_triggered
        )
        self.toolbar.save_trace_action.triggered.connect(
            self.on_save_found_lines_triggered
        )
        self.toolbar.clear_trace_action.triggered.connect(
            self.on_clear_found_lines_triggered
        )
        self.toolbar.configure_action.triggered.connect(
            self.on_configure_action_triggered
        )

        self.box_frequency.setup_ui_actions()
        self.box_frequency.changed.connect(self.on_frequency_box_changed)

        self.box_voltage.setup_ui_actions()
        self.box_voltage.changed.connect(self.on_voltage_box_changed)
        self.box_voltage.dataModeChanged.connect(self.on_voltage_box_data_mode_changed)

        self.box_find_lines.setup_ui_actions()
        self.box_find_lines.found_lines_changed.connect(
            self.on_automatically_found_lines_changed
        )
        self.box_find_lines.lines_found.connect(self.on_automatically_found_lines_found)
        self.box_find_lines.show_frequency_requested.connect(
            self.on_show_frequency_requested
        )

        self.box_found_lines.setup_ui_actions()
        self.box_found_lines.show_frequency_requested.connect(
            self.on_show_frequency_requested
        )
        self.box_found_lines.model.frequencies_removed.connect(
            self.on_table_rows_removed
        )

        line: pg.PlotDataItem
        for line in (self.automatically_found_lines, self.user_found_lines):
            line.sigPointsClicked.connect(self.on_points_clicked)

        self._view_all_action.triggered.connect(self.on_view_all_triggered)

        self.figure.sceneObj.sigMouseClicked.connect(self.on_plot_clicked)

    def on_lim_changed(self, arg: tuple[PlotWidget, list[list[float]]]) -> None:
        if self._ignore_scale_change.locked():
            return
        rect: list[list[float]] = arg[1]
        xlim: list[float]
        ylim: list[float]
        xlim, ylim = rect
        with self._ignore_scale_change:
            self.on_xlim_changed(xlim)
            self.on_ylim_changed(ylim)

    def on_xlim_changed(self, xlim: list[float]) -> None:
        min_freq, max_freq = min(xlim), max(xlim)
        self.box_frequency.set_range(min_freq, max_freq)
        self.set_x_range(*self.box_frequency.range)

    def on_ylim_changed(self, ylim: list[float | np.float64]) -> None:
        min_voltage, max_voltage = min(ylim), max(ylim)
        self.box_voltage.set_range(min_voltage, max_voltage)
        self.set_y_range(lower_value=min_voltage, upper_value=max_voltage)

    def set_x_range(
        self, lower_value: float | np.float64, upper_value: float | np.float64
    ) -> None:
        self.figure.getPlotItem().setXRange(lower_value, upper_value, padding=0.0)
        self.box_find_lines.current_freq = (upper_value + lower_value) / 2.0

    def set_y_range(
        self, lower_value: float | np.float64, upper_value: float | np.float64
    ) -> None:
        self.figure.getPlotItem().setYRange(lower_value, upper_value, padding=0.0)

    def ensure_y_fits(self) -> None:
        if (x := self._plot_line.xData) is None or x.size < 2:
            return
        if (y := self._plot_line.yData) is None or y.size < 2:
            return
        x_axis: pg.AxisItem = self._canvas.getAxis("bottom")
        y_axis: pg.AxisItem = self._canvas.getAxis("left")
        visible_points: NDArray[np.float64] = y[
            (x >= min(x_axis.range)) & (x <= max(x_axis.range))
        ]
        if np.any(visible_points < min(y_axis.range)):
            minimum: np.float64 = np.min(visible_points)
            # noinspection PyTypeChecker
            self.set_y_range(
                minimum - 0.05 * (max(y_axis.range) - minimum), max(y_axis.range)
            )
        if np.any(visible_points > max(y_axis.range)):
            maximum: np.float64 = np.max(visible_points)
            # noinspection PyTypeChecker
            self.set_y_range(
                min(y_axis.range), maximum + 0.05 * (maximum - min(y_axis.range))
            )

    @Slot(pg.PlotDataItem, np.ndarray, MouseClickEvent)
    def on_points_clicked(
        self, item: pg.PlotDataItem, points: Iterable[pg.SpotItem], ev: MouseClickEvent
    ) -> None:
        if item.xData is None or item.yData is None:
            return
        if not self.trace_mode:
            return
        if ev.button() != Qt.MouseButton.LeftButton:
            return

        point: pg.SpotItem
        if ev.modifiers() == Qt.KeyboardModifier.ShiftModifier:
            items: NDArray[np.float64] = item.scatter.data["item"]
            mask: NDArray[np.bool_] = np.full(items.shape, True, np.bool_)
            for point in points:
                mask &= items != point
                self.box_find_lines.remove_found_line(point.pos().x())
                self.user_found_lines_data = self.user_found_lines_data[
                    self.user_found_lines_data != point.pos().x()
                ]
                self.box_found_lines.model.remove_line(point.pos().x())
            item.setData(item.xData[mask], item.yData[mask])

            with the(not self.box_found_lines.model.is_empty) as enabled:
                self.toolbar.copy_trace_action.setEnabled(enabled)
                self.toolbar.save_trace_action.setEnabled(enabled)
                self.toolbar.clear_trace_action.setEnabled(enabled)

        elif ev.modifiers() == Qt.KeyboardModifier.NoModifier:
            found_lines_frequencies: NDArray[np.float64] = (
                self.box_found_lines.model.all_data(0)
            )
            selected_points: list[int] = [
                cast(int, np.argmin(np.abs(point.pos().x() - found_lines_frequencies)))
                for point in points
            ]
            self.box_found_lines.select(selected_points)

    @Slot(int)
    def on_automatically_found_lines_found(self, count: int) -> None:
        self.status_bar.showMessage(
            self.box_find_lines.tr("Found {} lines").format(count)
        )

    @Slot(frozenset)
    def on_table_rows_removed(self, frequencies: frozenset[float]) -> None:
        mask: NDArray[np.bool_] = np.full(
            self.user_found_lines_data.shape, True, np.bool_
        )
        for f in frequencies:
            mask[self.user_found_lines_data == f] = False
        self.user_found_lines_data = self.user_found_lines_data[mask]
        self.user_found_lines.setData(
            self.user_found_lines_data,
            self._plot_line.yData[
                self.box_found_lines.model.frequency_indices(
                    self._plot_data, self.user_found_lines_data
                )
            ],
        )

        for f in frequencies:
            self.box_find_lines.remove_found_line(f)
        self.automatically_found_lines.setData(
            self.box_find_lines.found_lines_freq,
            self._plot_line.yData[
                self.box_found_lines.model.frequency_indices(
                    self._plot_data, self.box_find_lines.found_lines_freq
                )
            ],
        )

        with the(not self.box_found_lines.model.is_empty) as enabled:
            self.toolbar.copy_trace_action.setEnabled(enabled)
            self.toolbar.save_trace_action.setEnabled(enabled)
            self.toolbar.clear_trace_action.setEnabled(enabled)

    @Slot(tuple)
    def on_mouse_moved(self, event: tuple[QPointF]) -> None:
        if self._plot_line.xData is None and self._plot_line.yData is None:
            return
        pos: QPointF = event[0]
        if self.figure.sceneBoundingRect().contains(pos):
            point: QPointF = self._canvas.getViewBox().mapSceneToView(pos)
            if self.figure.visibleRange().contains(point):
                self.status_bar.clearMessage()
                self._crosshair_v_line.setPos(point.x())
                self._crosshair_h_line.setPos(point.y())
                self._crosshair_h_line.setVisible(self.settings.show_crosshair)
                self._crosshair_v_line.setVisible(self.settings.show_crosshair)
                self._cursor_x.setVisible(True)
                self._cursor_y.setVisible(True)
                self._cursor_x.setValue(point.x())
                self._cursor_y.setValue(point.y())

                if self.settings.show_coordinates_at_crosshair:
                    self._cursor_balloon.setPos(point)
                    self._cursor_balloon.setHtml(
                        self._cursor_x.text() + "<br>" + self._cursor_y.text()
                    )
                    balloon_border: QRectF = self._cursor_balloon.boundingRect()
                    sx: float
                    sy: float
                    sx, sy = self._canvas.getViewBox().viewPixelSize()
                    balloon_width: float = balloon_border.width() * sx
                    balloon_height: float = balloon_border.height() * sy
                    anchor_x: float = (
                        0.0
                        if point.x() - self.figure.visibleRange().left() < balloon_width
                        else 1.0
                    )
                    anchor_y: float = (
                        0.0
                        if self.figure.visibleRange().bottom() - point.y()
                        < balloon_height
                        else 1.0
                    )
                    self._cursor_balloon.setAnchor((anchor_x, anchor_y))
                self._cursor_balloon.setVisible(
                    self.settings.show_coordinates_at_crosshair
                )
            else:
                self.hide_cursors()
        else:
            self.hide_cursors()

    @Slot()
    def on_view_all_triggered(self) -> None:
        self._canvas.getViewBox().autoRange(padding=0.0)

    @Slot(MouseClickEvent)
    def on_plot_clicked(self, event: MouseClickEvent) -> None:
        pos: QPointF = event.scenePos()
        if not self.trace_mode:
            return
        if (
            event.modifiers() != Qt.KeyboardModifier.NoModifier
            or event.button() != Qt.MouseButton.LeftButton
        ):
            return
        if not self.figure.sceneBoundingRect().contains(pos):
            return
        x_span: np.float64 = np.ptp(self._canvas.axes["bottom"]["item"].range)
        y_span: np.float64 = np.ptp(self._canvas.axes["left"]["item"].range)
        point: QPointF = self._canvas.getViewBox().mapSceneToView(pos)
        with the(self._plot_line.xData) as x, the(self._plot_line.yData) as y:
            if x is None or not x.size:
                return
            distance: NDArray[np.float64] = np.min(
                np.hypot(
                    (x - point.x()) / x_span,
                    (y - point.y()) / y_span,
                )
            )
            if distance > 0.01:
                return
            closest_point_index: np.int64 = np.argmin(
                np.hypot(
                    (x - point.x()) / x_span,
                    (y - point.y()) / y_span,
                )
            )

            x_point: float = x[closest_point_index]
            y_point: float = y[closest_point_index]

            # avoid the same point to be marked several times
            if (
                self.user_found_lines.xData is not None
                and self.user_found_lines.yData.size
                and np.any(
                    (self.user_found_lines.xData == x_point)
                    & (self.user_found_lines.yData == y_point)
                )
            ):
                return
            if (
                self.automatically_found_lines.xData is not None
                and self.automatically_found_lines.yData.size
                and np.any(
                    (self.automatically_found_lines.xData == x_point)
                    & (self.automatically_found_lines.yData == y_point)
                )
            ):
                return

            self.user_found_lines_data = np.append(self.user_found_lines_data, x_point)

            self.user_found_lines.setData(
                self.user_found_lines_data,
                y[
                    self.box_found_lines.model.frequency_indices(
                        self._plot_data, self.user_found_lines_data
                    )
                ],
            )

        self.box_found_lines.model.add_line(self._plot_data, x_point)
        if self.settings.copy_frequency:
            copy_to_clipboard(str(1e-6 * x_point))
        self.toolbar.copy_trace_action.setEnabled(True)
        self.toolbar.save_trace_action.setEnabled(True)
        self.toolbar.clear_trace_action.setEnabled(True)

    @Slot(float, float)
    def on_frequency_box_changed(self, min_freq: float, max_freq: float) -> None:
        if self._loading.locked():
            return
        with self._loading:
            self.set_x_range(
                lower_value=min_freq,
                upper_value=max_freq,
            )

    @Slot(float, float)
    def on_voltage_box_changed(self, min_y: float, max_y: float) -> None:
        if self._loading.locked():
            return
        with self._loading:
            self.set_y_range(
                lower_value=min_y,
                upper_value=max_y,
            )

    @Slot()
    def on_configure_action_triggered(self) -> None:
        preferences_dialog: Preferences = Preferences(self.settings, self)
        if preferences_dialog.exec() == Preferences.DialogCode.Rejected:
            return
        self._install_translation()
        self._setup_colors()
        self.set_plot_line_appearance()
        self.set_axis_line_appearance()
        self.set_marks_appearance()
        self.set_crosshair_lines_appearance()
        self.box_found_lines.model.fancy_table_numbers = (
            self.settings.fancy_table_numbers
        )
        self.box_found_lines.model.log10_gamma = self.settings.log10_gamma
        self.load_catalog()
        if self._data_mode == DataMode.PSK and self._plot_data.frequency_span > 0.0:
            jump: float = (
                round(self.settings.jump / self._plot_data.frequency_step)
                * self._plot_data.frequency_step
            )
            self.toolbar.differentiate_action.setEnabled(
                bool(0.0 < jump < 0.25 * self._plot_data.frequency_span)
            )
            if not (0.0 < jump < 0.25 * self._plot_data.frequency_span):
                self.toolbar.differentiate_action.blockSignals(True)
                self.toolbar.differentiate_action.setChecked(False)
                self.toolbar.differentiate_action.blockSignals(False)
        self.display_gamma_or_voltage()

    def hide_cursors(self) -> None:
        self._crosshair_h_line.setVisible(False)
        self._crosshair_v_line.setVisible(False)
        self._cursor_x.setVisible(False)
        self._cursor_y.setVisible(False)
        self._cursor_balloon.setVisible(False)

    @Slot()
    def on_open_action_triggered(self) -> None:
        loaded: bool = self.load_data()
        self.toolbar.load_trace_action.setEnabled(
            loaded or self.toolbar.load_trace_action.isEnabled()
        )
        if loaded:
            self.load_catalog()

    def load_catalog(self) -> None:
        if frozenset(self.box_found_lines.model.catalog_file_names) == frozenset(
            catalog_file_names := self.settings.catalog_paths
        ):
            return
        if not catalog_file_names:
            return
        with self.show_loading():
            label: str
            if len(catalog_file_names) > 1:
                label = tag(
                    "html",
                    "\n".join(
                        (
                            p_tag(self.tr("Loading catalogs:")),
                            tag(
                                "ul",
                                "\n".join(
                                    tag("li", str(fn)) for fn in catalog_file_names
                                ),
                            ),
                        )
                    ),
                )
            else:
                label = tag(
                    "html",
                    p_tag(
                        self.tr("Loading a catalog from<br>{}").format(
                            catalog_file_names[0]
                        )
                    ),
                )
            try:
                # noinspection PyPackageRequirements,PyUnusedImports
                from pycatsearch.catalog import Catalog
            except ImportError:
                self.status_bar.showMessage(
                    self.tr("Unable to load a catalog: Python package missing.")
                )
            else:
                from ..widgets.waiting_screen import WaitingScreen

                ws: WaitingScreen[Catalog] = WaitingScreen(
                    parent=self,
                    label=label,
                    target=Catalog,
                    args=catalog_file_names,
                    label_alignment=Qt.AlignmentFlag.AlignLeading,
                )
                cat: Catalog | None = ws.exec()
                if cat is None or cat.is_empty:
                    if ws.is_cancelled():
                        self.status_bar.showMessage(
                            self.tr("Loading has been cancelled.")
                        )
                    else:
                        self.status_bar.showMessage(
                            self.tr("Failed to load a catalog.")
                        )
                else:
                    self.status_bar.showMessage(self.tr("Catalogs loaded."))
                self.box_found_lines.model.catalog = (
                    cat or self.box_found_lines.model.catalog
                )

    @property
    def line(self) -> PlotDataItem:
        return self._plot_line

    @property
    def label(self) -> str | None:
        return self._plot_line.name()

    def set_plot_line_appearance(self) -> None:
        self._plot_line.setPen(
            pg.mkPen(self.settings.line_color, width=0.5 * self.settings.line_thickness)
        )
        self._ghost_line.setPen(
            pg.mkPen(
                self.settings.ghost_line_color, width=0.5 * self.settings.line_thickness
            )
        )
        self._canvas.replot()

    def set_axis_line_appearance(self) -> None:
        def _(axis: pg.AxisItem) -> None:
            styles: dict[QFont.Style, str] = {
                QFont.Style.StyleNormal: "normal",
                QFont.Style.StyleItalic: "italic",
                QFont.Style.StyleOblique: "oblique",
            }
            variants: dict[QFont.Capitalization, str] = {
                QFont.Capitalization.MixedCase: "normal",
                QFont.Capitalization.AllUppercase: "normal",  # see transforms below
                QFont.Capitalization.AllLowercase: "normal",  # see transforms below
                QFont.Capitalization.SmallCaps: "small-caps",
                QFont.Capitalization.Capitalize: "titling-caps",
            }
            transforms: dict[QFont.Capitalization, str] = {
                QFont.Capitalization.MixedCase: "none",
                QFont.Capitalization.AllUppercase: "uppercase",
                QFont.Capitalization.AllLowercase: "lowercase",
                QFont.Capitalization.SmallCaps: "none",  # see variants above
                QFont.Capitalization.Capitalize: "capitalize",
            }
            decorations: dict[tuple[bool, bool, bool], str] = {
                (False, False, False): "none",
                (False, False, True): "overline",
                (False, True, False): "underline",
                (False, True, True): "underline overline",
                (True, False, False): "line-through",
                (True, False, True): "overline line-through",
                (True, True, False): "underline line-through",
                (True, True, True): "underline overline line-through",
            }
            font: QFont = self.settings.axis_label_font
            axis.labelStyle.update(
                {
                    "font-family": font.family(),
                    "font-size": f"{font.pointSizeF()}pt",
                    "font-stretch": f"{font.stretch()}%",
                    "font-style": styles[font.style()],
                    "font-weight": f"{font.weight()}",
                    "font-variant-caps": variants[font.capitalization()],
                    "text-decoration-line": decorations[
                        (font.strikeOut(), font.underline(), font.overline())
                    ],
                    "text-transform": transforms[font.capitalization()],
                }
            )
            axis.setTickFont(self.settings.axis_tick_font)
            axis.setFont(self.settings.axis_label_font)
            pen: QPen = axis.pen()
            pen.setWidthF(self.settings.axis_thickness)
            axis.setPen(pen)

        _(self._canvas.getAxis("bottom"))
        _(self._canvas.getAxis("left"))

    def set_marks_appearance(self) -> None:
        pen: QPen = pg.mkPen(
            self.settings.mark_pen, width=0.5 * self.settings.mark_pen_thickness
        )
        brush: QBrush = pg.mkBrush(self.settings.mark_brush)
        self.automatically_found_lines.setSymbolPen(pen)
        self.automatically_found_lines.setSymbolBrush(brush)
        self.automatically_found_lines.setSymbolSize(self.settings.mark_size)
        self.user_found_lines.setSymbolPen(pen)
        self.user_found_lines.setSymbolBrush(brush)
        self.user_found_lines.setSymbolSize(self.settings.mark_size)
        self._canvas.replot()

    def set_crosshair_lines_appearance(self) -> None:
        pen: QPen = pg.mkPen(
            self.settings.crosshair_lines_color,
            width=0.5 * self.settings.crosshair_lines_thickness,
        )
        self._crosshair_v_line.setPen(pen)
        self._crosshair_h_line.setPen(pen)
        self._canvas.replot()

    @Slot(float)
    def on_show_frequency_requested(self, f: float) -> None:
        self.box_frequency.center = f
        self.ensure_y_fits()

    @Slot()
    def on_load_found_lines_triggered(self) -> None:
        def load_csv(fn: Path) -> Sequence[float]:
            sep: str = self.settings.csv_separator
            try:
                data: NDArray[np.float64] = (
                    np.loadtxt(
                        fn,
                        delimiter=sep,
                        usecols=(0,),
                        encoding="utf-8",
                        dtype=np.complex128,
                    ).real
                    * 1e6
                )
            except ValueError:
                return []
            else:
                return data[
                    (data >= self._plot_data.min_frequency)
                    & (data <= self._plot_data.max_frequency)
                ]

        # noinspection PyPackageRequirements
        def load_xlsx(fn: Path) -> Sequence[float]:
            from openpyxl.reader.excel import load_workbook  # type: ignore
            from openpyxl.workbook.workbook import Workbook  # type: ignore
            from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

            workbook: Workbook = load_workbook(
                fn, read_only=True, keep_vba=False, data_only=True
            )
            if len(workbook.sheetnames) != 1:
                return []
            sheet: Worksheet | None = workbook.active
            if sheet is None:
                return []

            data: list[float] = []
            reading_title: bool = True
            row: tuple[Any, ...]
            # noinspection PyUnresolvedReferences
            for row in sheet.values:
                if reading_title and isinstance(row[0], Number):
                    reading_title = False
                if not reading_title and not isinstance(row[0], Number):
                    break
                if not reading_title and isinstance(row[0], Number):
                    data.append(float(row[0]))
            if not data:
                return []

            # noinspection PyTypeChecker
            data_: NDArray[np.float64] = np.asarray(data, dtype=np.float64) * 1e6
            return data_[
                (data_ >= self._plot_data.min_frequency)
                & (data_ <= self._plot_data.max_frequency)
            ]

        mimetypes.init()

        supported_formats_callbacks: dict[str, Callable[[Path], Sequence[float]]] = {
            mimetypes.types_map[".csv"]: load_csv,
            mimetypes.types_map[".txt"]: load_csv,
        }
        if importlib.util.find_spec("openpyxl") is not None:
            supported_formats_callbacks[mimetypes.types_map[".xlsx"]] = load_xlsx

        if not (filename := self._open_table_dialog.get_open_filename()):
            return

        file_type: str | None = mimetypes.guess_type(filename)[0]
        if file_type is None or file_type not in supported_formats_callbacks:
            return
        new_lines: Sequence[float] = supported_formats_callbacks[file_type](filename)
        if not len(new_lines):
            return

        self.box_found_lines.model.add_lines(self._plot_data, new_lines)
        # add the new lines to the marked ones
        self.user_found_lines_data = np.concatenate(
            (self.user_found_lines_data, new_lines)
        )
        # avoid duplicates
        self.user_found_lines_data = self.user_found_lines_data[
            np.unique(self.user_found_lines_data, return_index=True)[1]
        ]
        # plot the data
        self.user_found_lines.setData(
            self.user_found_lines_data,
            self._plot_line.yData[
                self.box_found_lines.model.frequency_indices(
                    self._plot_data, self.user_found_lines_data
                )
            ],
        )
        self.toolbar.copy_trace_action.setEnabled(True)
        self.toolbar.save_trace_action.setEnabled(True)
        self.toolbar.clear_trace_action.setEnabled(True)

    @Slot()
    def on_copy_found_lines_triggered(self) -> None:
        with self.show_loading():
            copy_to_clipboard(
                self.box_found_lines.table.stringify_table_plain_text(),
                self.box_found_lines.table.stringify_table_html(),
                Qt.TextFormat.RichText,
            )

    @Slot()
    def on_save_found_lines_triggered(self) -> None:
        def save_csv(fn: Path) -> None:
            from ..utils import remove_html

            sep: str = self.settings.csv_separator
            with (
                open(fn, "w", encoding="utf-8") as f_out,
                the(self.box_found_lines.model.header) as header,
            ):
                f_out.writelines(
                    map(
                        lambda s: "# " + s + "\n",
                        [
                            sep.join(
                                h.name if isinstance(h, HeaderWithUnit) else h
                                for h in header
                            ),
                            sep.join(
                                h.unit if isinstance(h, HeaderWithUnit) else ""
                                for h in header
                            ),
                        ],
                    )
                )
                for row in data:
                    f_out.write(remove_html(sep.join(map(str, row))) + "\n")

        def save_rtf(fn: Path) -> None:
            from ..utils import html_to_rtf, tag

            with (
                open(fn, "w", encoding="utf-8") as f_out,
            ):
                f_out.write(
                    html_to_rtf(
                        tag(
                            "html",
                            self.box_found_lines.table.stringify_table_html(
                                whole_table=True, with_headers=True
                            ),
                        )
                    )
                )

        def save_xlsx(fn: Path) -> None:
            with pd.ExcelWriter(fn) as writer:
                df: pd.DataFrame = pd.DataFrame(data)
                df.to_excel(
                    writer,
                    index=False,
                    header=list(map(str, self.box_found_lines.model.header)),
                    sheet_name=_translate("workbook", "Sheet1"),
                )

        supported_formats_callbacks: dict[str, Callable[[Path], None]] = {
            ".csv": save_csv,
            ".rtf": save_rtf,
        }
        if importlib.util.find_spec("openpyxl") is not None:
            supported_formats_callbacks[".xlsx"] = save_xlsx

        if not (filename := self._save_table_dialog.get_save_filename()):
            return

        data: list[list[object]] = [
            [
                (
                    item * 1e-6
                    if isinstance(
                        item := self.box_found_lines.model.item(row, column),
                        float,
                    )
                    and column == 0
                    else item
                )
                for column in range(self.box_found_lines.model.columnCount())
            ]
            for row in range(self.box_found_lines.model.rowCount(available_count=True))
        ]

        filename_ext: str = filename.suffix.casefold()
        if filename_ext in supported_formats_callbacks:
            with self.show_loading():
                supported_formats_callbacks[filename_ext](filename)

    @Slot(np.ndarray)
    def on_automatically_found_lines_changed(self, freq: NDArray[np.double]) -> None:
        self.automatically_found_lines.setData(
            freq,
            self._plot_data.y_data[
                self.box_found_lines.model.frequency_indices(
                    self._plot_data, self.box_find_lines.found_lines_freq
                )
            ],
        )
        self._canvas.replot()

        self.box_found_lines.model.set_lines(
            self._plot_data, self.user_found_lines_data, freq
        )
        with the(not self.box_found_lines.model.is_empty) as enabled:
            self.toolbar.copy_trace_action.setEnabled(enabled)
            self.toolbar.save_trace_action.setEnabled(enabled)
            self.toolbar.clear_trace_action.setEnabled(enabled)

    @Slot()
    def on_clear_found_lines_triggered(self) -> None:
        self.clear_found_lines()

    def clear_found_lines(self) -> None:
        self.automatically_found_lines.clear()
        self.box_find_lines.blockSignals(True)
        self.box_find_lines.clear_found_lines()
        self.box_find_lines.blockSignals(False)
        self.user_found_lines.clear()
        self.user_found_lines_data = np.empty(0)
        self.box_found_lines.model.clear()
        self.toolbar.copy_trace_action.setEnabled(False)
        self.toolbar.save_trace_action.setEnabled(False)
        self.toolbar.clear_trace_action.setEnabled(False)
        self._canvas.replot()

    @Slot()
    def on_clear_action_triggered(self) -> None:
        close: QMessageBox = QMessageBox()
        close.setText(self.tr("Are you sure?"))
        close.setIcon(QMessageBox.Icon.Question)
        close.setWindowIcon(self.windowIcon())
        close.setWindowTitle(self.tr("Spectrometer Data Viewer"))
        close.setStandardButtons(
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
        )
        if close.exec() != QMessageBox.StandardButton.Yes:
            return

        self._ghost_line.clear()
        self._plot_line.clear()
        self._ghost_data.clear()
        self._plot_data.clear()
        self.clear_found_lines()
        self.toolbar.trace_action.setChecked(False)
        self.toolbar.clear_action.setEnabled(False)
        self.toolbar.open_ghost_action.setEnabled(False)
        self.toolbar.clear_ghost_action.setEnabled(False)
        self.toolbar.differentiate_action.setEnabled(False)
        self.toolbar.save_data_action.setEnabled(False)
        self.toolbar.copy_figure_action.setEnabled(False)
        self.toolbar.save_figure_action.setEnabled(False)
        self.toolbar.trace_action.setEnabled(False)
        self.toolbar.load_trace_action.setEnabled(False)
        self.toolbar.copy_trace_action.setEnabled(False)
        self.toolbar.save_trace_action.setEnabled(False)
        self.toolbar.clear_trace_action.setEnabled(False)
        self._cursor_balloon.setVisible(False)
        self._crosshair_h_line.setVisible(False)
        self._crosshair_v_line.setVisible(False)
        self._cursor_x.setVisible(True)
        self._cursor_y.setVisible(True)
        self._canvas.replot()
        self.status_bar.clearMessage()
        self.setWindowTitle(self.tr("Spectrometer Data Viewer"))
        self._data_mode = DataMode.unknown

    @Slot()
    def on_open_ghost_action_triggered(self) -> None:
        self.load_ghost_data()

    @Slot()
    def on_clear_ghost_action_triggered(self) -> None:
        self.clear_ghost()

    def clear_ghost(self) -> None:
        self._ghost_line.clear()
        self._ghost_data.clear()
        self.toolbar.clear_ghost_action.setEnabled(False)
        self._canvas.replot()

    def configure_interface_after_loading_data(self, f: NDArray[np.float64]) -> None:
        if self._data_mode in (DataMode.FS,):
            self.box_voltage.show_gamma = False

        self.toolbar.clear_action.setEnabled(True)

        min_frequency: np.float64 = cast(np.float64, f[0])
        max_frequency: np.float64 = cast(np.float64, f[-1])

        self.box_frequency.switch_range(min_frequency, max_frequency)

        step: int = int(
            round(self.settings.jump / ((max_frequency - min_frequency) / (f.size - 1)))
        )
        self.toolbar.differentiate_action.setEnabled(
            self._data_mode == DataMode.PSK and 0 < step < 0.25 * f.size
        )

        self.box_voltage.can_show_gamma = self._data_mode in (
            DataMode.PSK,
            DataMode.PSK_WITH_JUMP,
            DataMode.TIME_DOMAIN,
        )
        self.toolbar.save_data_action.setEnabled(True)
        self.toolbar.copy_figure_action.setEnabled(True)
        self.toolbar.save_figure_action.setEnabled(True)
        self.toolbar.trace_action.setEnabled(True)
        self.toolbar.load_trace_action.setEnabled(True)

    def load_data(self, filename: Path | None = None) -> bool:
        self.clear_ghost()
        self.clear_found_lines()

        if not filename and not (
            filename := self._open_data_dialog.get_open_filename()
        ):
            return False

        data: SpectrometerData | None = load_data(self, filename)
        if data is None:
            return False
        if self._data_mode == data.mode:
            return self.set_data(data)

        if data.mode in FrequencyDomainWindow.supported_modes:
            # noinspection PyTypeChecker
            w = FrequencyDomainWindow(parent=self.parent(), flags=self.windowFlags())
            r: bool = w.set_data(data)
            if r:
                if self.box_found_lines.model.catalog is None:
                    self.load_catalog()
                w.box_found_lines.model.catalog = self.box_found_lines.model.catalog
                w.show()
                if self._data_mode == DataMode.unknown:
                    self.close()
            else:
                w.deleteLater()
            return r

        from . import TimeDomainWindow

        if data.mode in TimeDomainWindow.supported_modes:
            # noinspection PyTypeChecker
            w = TimeDomainWindow(parent=self.parent(), flags=self.windowFlags())
            r: bool = w.set_data(data)
            if r:
                w.show()
                if self._data_mode == DataMode.unknown:
                    self.close()
            else:
                w.deleteLater()
            return r

        return False

    def set_data(self, data: SpectrometerData | None) -> bool:
        if data is None:
            return False

        filename: Path
        v: NDArray[np.float64]
        f: NDArray[np.float64]
        g: NDArray[np.float64]
        t: NDArray[np.float64]
        data_mode: DataMode
        filename, f, g, v, t, data_mode = data

        if data_mode not in FrequencyDomainWindow.supported_modes:
            return False

        if not (f.size and v.size):
            return False

        self.settings.display_processing = data_mode in (
            DataMode.PSK,
            DataMode.PSK_WITH_JUMP,
        )
        self._data_mode = data_mode

        self._plot_data.set_data(
            frequency_data=f, gamma_data=g, voltage_data=v, time_data=t
        )
        self.box_find_lines.set_spectrum(f, v, g, self._data_mode)
        self.configure_interface_after_loading_data(f)

        self._plot_data.x_data_type = PlotDataItem.FREQUENCY_DATA
        self._ghost_data.x_data_type = PlotDataItem.FREQUENCY_DATA

        self._plot_line.setData(
            name=str(filename.parent / filename.stem),
        )

        self.display_gamma_or_voltage()

        self.set_x_range(*self.box_frequency.range)
        self.set_y_range(*self.box_voltage.range)

        self.setWindowTitle(self.tr("%s — Spectrometer Data Viewer") % filename)

        self.toolbar.open_ghost_action.setEnabled(True)

        return True

    def load_ghost_data(self, filename: Path | None = None) -> bool:
        if not filename and not (
            filename := self._open_data_dialog.get_open_filename()
        ):
            return False

        data: SpectrometerData | None = load_data(self, filename)
        if data is None:
            return False

        v: NDArray[np.float64]
        f: NDArray[np.float64]
        g: NDArray[np.float64]
        t: NDArray[np.float64]
        data_mode: DataMode
        filename, f, g, v, t, data_mode = data

        if data_mode not in FrequencyDomainWindow.supported_modes:
            return False

        if data_mode != self._data_mode:
            return False

        if not (f.size and v.size):
            return False

        self._ghost_data.set_data(
            frequency_data=f, gamma_data=g, voltage_data=v, time_data=t
        )
        self._ghost_line.setData(
            name=str(filename.parent / filename.stem),
        )

        self.toolbar.clear_ghost_action.setEnabled(True)

        self.display_gamma_or_voltage()

        return True

    @property
    def trace_mode(self) -> bool:
        return self.toolbar.trace_action.isChecked()

    def actions_off(self) -> None:
        self.toolbar.trace_action.setChecked(False)

    @Slot(bool)
    def on_differentiate_action_toggled(self, on: bool) -> None:
        self._data_mode = DataMode.PSK_WITH_JUMP if on else DataMode.PSK
        self.display_gamma_or_voltage()
        self.box_find_lines.set_spectrum(
            self._plot_data.frequency_data,
            self._plot_data.voltage_data,
            self._plot_data.gamma_data,
            self._data_mode,
        )
        self.box_found_lines.model.refresh(self._plot_data)

    @Slot(str)
    def on_voltage_box_data_mode_changed(self, mode: str) -> None:
        self._plot_data.y_data_type = mode
        self._ghost_data.y_data_type = mode
        self.box_find_lines.set_data_type(mode)
        self.display_gamma_or_voltage()

    def setup_left_axis(self) -> None:
        a: pg.AxisItem = self._canvas.getAxis("left")
        if self._plot_data.y_data_type == PlotDataItem.GAMMA_DATA:
            a.enableAutoSIPrefix(False)
            a.setLabel(
                text=_translate("plot axes labels", "Absorption"),
                units=_translate("unit", "cm<sup>−1</sup>"),
            )
            a.scale = 1.0
            a.autoSIPrefixScale = 1.0

            self._cursor_y.suffix = _translate("unit", "cm<sup>−1</sup>")
            self._cursor_y.siPrefix = False
            self._cursor_y.setFormatStr(
                "{mantissa:.{decimals}f}×10<sup>{exp}</sup>{suffixGap}{suffix}"
            )

        elif self._plot_data.y_data_type == PlotDataItem.VOLTAGE_DATA:
            a.enableAutoSIPrefix(True)
            a.setLabel(
                text=_translate("plot axes labels", "Voltage"),
                units=_translate("unit", "V"),
            )

            self._cursor_y.suffix = _translate("unit", "V")
            self._cursor_y.siPrefix = True
            self._cursor_y.setFormatStr(
                "{scaledValue:.{decimals}f}{suffixGap}{siPrefix}{suffix}"
            )

        else:
            raise ValueError(f"Invalid data type: {self._plot_data.y_data_type!r}")

    def display_gamma_or_voltage(self) -> None:
        if self._data_mode == DataMode.PSK_WITH_JUMP:
            self._plot_data.jump = self.settings.jump
            self._ghost_data.jump = self.settings.jump
        else:
            self._plot_data.jump = np.nan
            self._ghost_data.jump = np.nan

        if self._plot_data:  # something is loaded
            self._plot_line.setData(self._plot_data.x_data, self._plot_data.y_data)

            y_data: NDArray[np.float64] = self._plot_data.y_data
            min_y: np.float64 = np.min(y_data)
            max_y: np.float64 = np.max(y_data)
            with self._loading:
                self.box_voltage.switch_range(min_y, max_y)
            self.on_ylim_changed([min_y, max_y])

        if self._ghost_data:  # something is loaded
            self._ghost_line.setData(self._ghost_data.x_data, self._ghost_data.y_data)

        if self.box_find_lines.found_lines_freq.size:  # something is marked
            self.automatically_found_lines.setData(
                self.box_find_lines.found_lines_freq,
                self._plot_data.y_data[
                    self.box_found_lines.model.frequency_indices(
                        self._plot_data, self.box_find_lines.found_lines_freq
                    )
                ],
            )
        if self.user_found_lines_data.size:  # something is marked
            self.user_found_lines.setData(
                self.user_found_lines_data,
                self._plot_data.y_data[
                    self.box_found_lines.model.frequency_indices(
                        self._plot_data, self.user_found_lines_data
                    )
                ],
            )

        self.setup_left_axis()
        self.hide_cursors()

    @Slot()
    def on_save_data_triggered(self) -> None:
        if self._plot_line.yData is None:
            return

        def save_csv(fn: Path) -> None:
            data: NDArray[np.float64]
            sep: str = self.settings.csv_separator
            if self.box_voltage.show_gamma:
                data = np.column_stack((x * 1e-6, y))
                # noinspection PyTypeChecker
                np.savetxt(
                    fn,
                    data,
                    delimiter=sep,
                    header=(
                        sep.join(
                            (
                                _translate("plot axes labels", "Frequency"),
                                _translate("plot axes labels", "Absorption"),
                            )
                        )
                        + "\n"
                        + sep.join(
                            (
                                _translate("unit", "MHz"),
                                _translate("unit", "cm⁻¹"),
                            )
                        )
                    ),
                    fmt=("%.3f", "%.6e"),
                    encoding="utf-8",
                )
            else:
                data = np.column_stack((x * 1e-6, y * 1e3))
                # noinspection PyTypeChecker
                np.savetxt(
                    fn,
                    data,
                    delimiter=sep,
                    header=(
                        sep.join(
                            (
                                _translate("plot axes labels", "Frequency"),
                                _translate("plot axes labels", "Voltage"),
                            )
                        )
                        + "\n"
                        + sep.join(
                            (
                                _translate("unit", "MHz"),
                                _translate("unit", "mV"),
                            )
                        )
                    ),
                    fmt=("%.3f", "%.6f"),
                    encoding="utf-8",
                )

        def save_rtf(fn: Path) -> None:
            from ..utils import html_to_rtf, tag

            table: list[list[str]] = []
            if self.box_voltage.show_gamma:
                table.append(
                    list(
                        map(
                            str,
                            [
                                self.box_found_lines.model.header[0],
                                self.box_found_lines.model.header[2],
                            ],
                        )
                    )
                )
                for _x, _y in zip(x, y, strict=True):
                    table.append([f"{_x:.3f}", f"{_y:.6e}"])
            else:
                table.append(
                    list(
                        map(
                            str,
                            [
                                self.box_found_lines.model.header[0],
                                self.box_found_lines.model.header[1],
                            ],
                        )
                    )
                )
                for _x, _y in zip(x, y * 1e3, strict=True):
                    table.append([f"{_x:.3f}", f"{_y:.6f}"])
            with open(fn, "w", encoding="utf-8") as f_out:
                f_out.write(
                    html_to_rtf(
                        tag(
                            "html",
                            tag(
                                "table",
                                "".join(
                                    tag(
                                        "tr",
                                        "".join(tag("td", cell) for cell in row),
                                    )
                                    for row in table
                                ),
                            ),
                        )
                    )
                )

        def save_xlsx(fn: Path) -> None:
            data: NDArray[np.float64]
            with pd.ExcelWriter(fn) as writer:
                df: pd.DataFrame
                if self.box_voltage.show_gamma:
                    data = np.column_stack((x * 1e-6, y))
                    df = pd.DataFrame(data)
                    df.to_excel(
                        writer,
                        index=False,
                        header=list(
                            map(
                                str,
                                [
                                    self.box_found_lines.model.header[0],
                                    self.box_found_lines.model.header[2],
                                ],
                            )
                        ),
                        sheet_name=self._plot_line.name()
                        or _translate("workbook", "Sheet1"),
                    )
                else:
                    data = np.column_stack((x * 1e-6, y * 1e3))
                    df = pd.DataFrame(data)
                    df.to_excel(
                        writer,
                        index=False,
                        header=list(
                            map(
                                str,
                                [
                                    self.box_found_lines.model.header[0],
                                    self.box_found_lines.model.header[1],
                                ],
                            )
                        ),
                        sheet_name=self._plot_line.name()
                        or _translate("workbook", "Sheet1"),
                    )

        supported_formats_callbacks: dict[str, Callable[[Path], None]] = {
            ".csv": save_csv,
            ".rtf": save_rtf,
            ".xlsx": save_xlsx,
        }

        if not (filename := self._save_table_dialog.get_save_filename()):
            return
        x: NDArray[np.float64] = self._plot_line.xData
        y: NDArray[np.float64] = self._plot_line.yData
        max_mark: float
        min_mark: float
        min_mark, max_mark = self._canvas.axes["bottom"]["item"].range
        good: NDArray[np.bool_] = (min_mark <= x) & (x <= max_mark)
        x = x[good]
        y = y[good]
        del good

        filename_ext: str = filename.suffix.casefold()
        if filename_ext in supported_formats_callbacks:
            with self.show_loading():
                supported_formats_callbacks[filename_ext](filename)

    @Slot()
    def on_copy_figure_triggered(self) -> None:
        exporter: ImageExporter = ImageExporter(self._canvas)
        self.hide_cursors()
        exporter.export(copy=True)

    @Slot()
    def on_save_figure_triggered(self) -> None:
        exporter: ImageExporter = ImageExporter(self._canvas)
        if not (filename := self._save_image_dialog.get_save_filename()):
            return
        self.hide_cursors()
        exporter.export(str(filename))
